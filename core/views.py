from django.http import HttpResponse
from collections import defaultdict

from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render

from .forms import (
    DeliveryItemForm,
    DeliveryNoteForm,
    InvoiceForm,
    InvoiceItemForm,
    StockMovementForm,
)
from .models import DeliveryItem, DeliveryNote, Invoice, InvoiceItem, StockItem, StockMovement


@login_required
def dashboard(request):
    return render(request, 'core/dashboard.html', {
        'invoices': Invoice.objects.count(),
        'deliveries': DeliveryNote.objects.count(),
        'stock_items': StockItem.objects.count(),
        'open_invoices': Invoice.objects.filter(status='open').count(),
        'partial_invoices': Invoice.objects.filter(status='partial').count(),
        'complete_invoices': Invoice.objects.filter(status='complete').count(),
    })


@login_required
def invoice_list(request):
    invoices = Invoice.objects.all().order_by('-created_at')

    search = request.GET.get('search', '').strip()
    supplier = request.GET.get('supplier', '').strip()
    status = request.GET.get('status', '').strip()

    if search:
        invoices = invoices.filter(invoice_number__icontains=search)

    if supplier:
        invoices = invoices.filter(supplier_name__icontains=supplier)

    if status:
        invoices = invoices.filter(status=status)

    return render(request, 'core/invoice_list.html', {
        'invoices': invoices,
        'search': search,
        'supplier': supplier,
        'status': status,
    })


@login_required
def invoice_create(request):
    if request.method == 'POST':
        form = InvoiceForm(request.POST, request.FILES)
        if form.is_valid():
            invoice = form.save(commit=False)
            invoice.created_by = request.user
            invoice.save()
            return redirect('invoice_list')
    else:
        form = InvoiceForm()
    return render(request, 'core/form.html', {'form': form, 'title': 'Rechnung erstellen'})


@login_required
def invoice_item_add(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    if request.method == 'POST':
        form = InvoiceItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.invoice = invoice
            item.save()
            return redirect('invoice_list')
    else:
        form = InvoiceItemForm()
    return render(request, 'core/form.html', {
        'form': form,
        'title': f'Artikel zur Rechnung hinzufügen {invoice.invoice_number}'
    })


@login_required
def delivery_list(request):
    deliveries = DeliveryNote.objects.all().order_by('-created_at')

    search = request.GET.get('search', '').strip()
    supplier = request.GET.get('supplier', '').strip()

    if search:
        deliveries = deliveries.filter(delivery_number__icontains=search)

    if supplier:
        deliveries = deliveries.filter(supplier_name__icontains=supplier)

    return render(request, 'core/delivery_list.html', {
        'deliveries': deliveries,
        'search': search,
        'supplier': supplier,
    })


@login_required
def delivery_create(request):
    if request.method == 'POST':
        form = DeliveryNoteForm(request.POST, request.FILES)
        if form.is_valid():
            delivery = form.save(commit=False)
            delivery.created_by = request.user
            delivery.save()
            return redirect('delivery_list')
    else:
        form = DeliveryNoteForm()
    return render(request, 'core/form.html', {'form': form, 'title': 'Lieferschein erstellen'})


@login_required
def delivery_item_add(request, delivery_id):
    delivery = get_object_or_404(DeliveryNote, id=delivery_id)
    if request.method == 'POST':
        form = DeliveryItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.delivery_note = delivery
            item.save()
            return redirect('delivery_list')
    else:
        form = DeliveryItemForm()
    return render(request, 'core/form.html', {
        'form': form,
        'title': f'Artikel zur Lieferung hinzufügen {delivery.delivery_number}'
    })


@login_required
def compare_documents(request):
    invoices = Invoice.objects.prefetch_related('items', 'delivery_notes__items').all()
    grouped_results = []

    for invoice in invoices:
        delivered_map = defaultdict(int)
        linked_deliveries = list(invoice.delivery_notes.all())

        for delivery_note in linked_deliveries:
            for delivery_item in delivery_note.items.all():
                key = delivery_item.item_name.strip().lower()
                delivered_map[key] += delivery_item.quantity

        invoice_status = 'complete' if invoice.items.exists() else 'open'
        item_results = []
        total_expected = 0
        total_delivered = 0

        for invoice_item in invoice.items.all():
            key = invoice_item.item_name.strip().lower()
            delivered_qty = delivered_map.get(key, 0)
            difference = delivered_qty - invoice_item.quantity

            total_expected += invoice_item.quantity
            total_delivered += delivered_qty

            if delivered_qty == 0:
                status = 'Offen'
                invoice_status = 'open'
            elif delivered_qty < invoice_item.quantity:
                status = 'In Bearbeitung'
                invoice_status = 'partial'
            elif delivered_qty == invoice_item.quantity:
                status = 'Abgeschlossen'
                if invoice_status not in ['open', 'partial', 'difference']:
                    invoice_status = 'complete'
            else:
                status = 'Differenz'
                invoice_status = 'difference'

            item_results.append({
                'item_name': invoice_item.item_name,
                'expected_quantity': invoice_item.quantity,
                'delivered_quantity': delivered_qty,
                'difference': difference,
                'status': status,
            })

        invoice.status = invoice_status
        invoice.save()

        grouped_results.append({
            'invoice': invoice,
            'deliveries': linked_deliveries,
            'items': item_results,
            'total_expected': total_expected,
            'total_delivered': total_delivered,
            'total_difference': total_delivered - total_expected,
        })

    return render(request, 'core/compare.html', {'grouped_results': grouped_results})


@login_required
def stock_list(request):
    stock_items = StockItem.objects.all().order_by('item_name')

    search = request.GET.get('search', '').strip()
    if search:
        stock_items = stock_items.filter(item_name__icontains=search)

    return render(request, 'core/stock_list.html', {
        'stock_items': stock_items,
        'search': search,
    })


@login_required
def apply_stock_from_deliveries(request):
    delivery_items = DeliveryItem.objects.all()
    aggregated = defaultdict(int)

    for item in delivery_items:
        aggregated[item.item_name.strip()] += item.quantity

    for item_name, quantity in aggregated.items():
        stock_item, _ = StockItem.objects.get_or_create(item_name=item_name)
        change = quantity - stock_item.current_quantity

        if change != 0:
            stock_item.current_quantity = quantity
            stock_item.save()

            StockMovement.objects.create(
                stock_item=stock_item,
                movement_type='goods_receipt',
                quantity_change=change,
                reason='Automatic stock update from deliveries',
                created_by=request.user
            )

    return redirect('stock_list')


@login_required
def stock_movement_create(request):
    if request.method == 'POST':
        form = StockMovementForm(request.POST)
        if form.is_valid():
            movement = form.save(commit=False)
            movement.created_by = request.user
            movement.save()

            stock_item = movement.stock_item
            stock_item.current_quantity += movement.quantity_change
            stock_item.save()

            return redirect('stock_list')
    else:
        form = StockMovementForm()
    return render(request, 'core/form.html', {'form': form, 'title': 'Lagerbestand bearbeiten'})


@login_required
def invoice_detail(request, invoice_id):
    invoice = get_object_or_404(
        Invoice.objects.prefetch_related('items', 'delivery_notes__items'),
        id=invoice_id
    )
    return render(request, 'core/invoice_detail.html', {'invoice': invoice})


@login_required
def delivery_detail(request, delivery_id):
    delivery = get_object_or_404(
        DeliveryNote.objects.prefetch_related('items'),
        id=delivery_id
    )
    return render(request, 'core/delivery_detail.html', {'delivery': delivery})


@login_required
def stock_movement_list(request):
    movements = StockMovement.objects.select_related('stock_item', 'created_by').order_by('-created_at')
    return render(request, 'core/stock_movement_list.html', {'movements': movements})


@login_required
def invoice_item_edit(request, item_id):
    item = get_object_or_404(InvoiceItem, id=item_id)
    if request.method == 'POST':
        form = InvoiceItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect('invoice_detail', invoice_id=item.invoice.id)
    else:
        form = InvoiceItemForm(instance=item)

    return render(request, 'core/form.html', {
        'form': form,
        'title': f'Rechnungsposition bearbeiten: {item.item_name}'
    })


@login_required
def invoice_item_delete(request, item_id):
    item = get_object_or_404(InvoiceItem, id=item_id)
    invoice_id = item.invoice.id

    if request.method == 'POST':
        item.delete()
        return redirect('invoice_detail', invoice_id=invoice_id)

    return render(request, 'core/confirm_delete.html', {
        'title': 'Rechnungsartikel löschen',
        'message': f'Sind Sie sicher, dass Sie "{item.item_name}" aus der {item.invoice.invoice_number} löschen möchten?'
    })


@login_required
def delivery_item_edit(request, item_id):
    item = get_object_or_404(DeliveryItem, id=item_id)
    if request.method == 'POST':
        form = DeliveryItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect('delivery_detail', delivery_id=item.delivery_note.id)
    else:
        form = DeliveryItemForm(instance=item)

    return render(request, 'core/form.html', {
        'form': form,
        'title': f'Lieferartikel bearbeiten: {item.item_name}'
    })


@login_required
def delivery_item_delete(request, item_id):
    item = get_object_or_404(DeliveryItem, id=item_id)
    delivery_id = item.delivery_note.id

    if request.method == 'POST':
        item.delete()
        return redirect('delivery_detail', delivery_id=delivery_id)

    return render(request, 'core/confirm_delete.html', {
        'title': 'Lieferartikel löschen',
        'message': f'Möchten Sie diese Sendung wirklich entfernen "{item.item_name}" aus der Zustellung {item.delivery_note.delivery_number}?'
    })


@login_required
def invoice_edit(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    if request.method == 'POST':
        form = InvoiceForm(request.POST, request.FILES, instance=invoice)
        if form.is_valid():
            updated_invoice = form.save(commit=False)
            updated_invoice.created_by = invoice.created_by
            updated_invoice.save()
            return redirect('invoice_detail', invoice_id=invoice.id)
    else:
        form = InvoiceForm(instance=invoice)

    return render(request, 'core/form.html', {
        'form': form,
        'title': f'Rechnung bearbeiten {invoice.invoice_number}'
    })


@login_required
def invoice_delete(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)

    if request.method == 'POST':
        invoice.delete()
        return redirect('invoice_list')

    return render(request, 'core/confirm_delete.html', {
        'title': 'Rechnung löschen',
        'message': f'Sind Sie sicher, dass Sie die Rechnung <span class="text-danger fw-bold fs-6">"{invoice.invoice_number}"</span> löschen möchten? Alle verknüpften Rechnungspositionen und Lieferscheine werden ebenfalls gelöscht.'
    })


@login_required
def delivery_edit(request, delivery_id):
    delivery = get_object_or_404(DeliveryNote, id=delivery_id)
    if request.method == 'POST':
        form = DeliveryNoteForm(request.POST, request.FILES, instance=delivery)
        if form.is_valid():
            updated_delivery = form.save(commit=False)
            updated_delivery.created_by = delivery.created_by
            updated_delivery.save()
            return redirect('delivery_detail', delivery_id=delivery.id)
    else:
        form = DeliveryNoteForm(instance=delivery)

    return render(request, 'core/form.html', {
        'form': form,
        'title': f'Lieferschein bearbeiten {delivery.delivery_number}'
    })


@login_required
def delivery_delete(request, delivery_id):
    delivery = get_object_or_404(DeliveryNote, id=delivery_id)

    if request.method == 'POST':
        delivery.delete()
        return redirect('delivery_list')

    return render(request, 'core/confirm_delete.html', {
        'title': 'Lieferschein löschen',
        'message': f'Sind Sie sicher, dass Sie den Lieferschein <span class="text-danger fw-bold fs-6">"{delivery.delivery_number}"</span> löschen möchten?'
    })


@login_required
def dashboard(request):
    recent_invoices = Invoice.objects.order_by('-created_at')[:5]
    recent_deliveries = DeliveryNote.objects.order_by('-created_at')[:5]
    recent_movements = StockMovement.objects.select_related('stock_item').order_by('-created_at')[:5]

    return render(request, 'core/dashboard.html', {
        'invoices': Invoice.objects.count(),
        'deliveries': DeliveryNote.objects.count(),
        'stock_items': StockItem.objects.count(),
        'open_invoices': Invoice.objects.filter(status='open').count(),
        'partial_invoices': Invoice.objects.filter(status='partial').count(),
        'complete_invoices': Invoice.objects.filter(status='complete').count(),
        'recent_invoices': recent_invoices,
        'recent_deliveries': recent_deliveries,
        'recent_movements': recent_movements,
    })


@login_required
def stock_movement_edit(request, movement_id):
    movement = get_object_or_404(StockMovement, id=movement_id)

    if request.method == 'POST':
        form = StockMovementForm(request.POST, instance=movement)
        if form.is_valid():
            old_quantity = movement.quantity_change

            updated = form.save(commit=False)
            updated.created_by = movement.created_by
            updated.save()

            # adjust stock difference
            stock_item = updated.stock_item
            stock_item.current_quantity += (updated.quantity_change - old_quantity)
            stock_item.save()

            return redirect('stock_movement_list')
    else:
        form = StockMovementForm(instance=movement)

    return render(request, 'core/form.html', {
        'form': form,
        'title': f'Lagerbestand bearbeiten'
    })


@login_required
def stock_movement_delete(request, movement_id):
    movement = get_object_or_404(StockMovement, id=movement_id)

    if request.method == 'POST':
        stock_item = movement.stock_item

        # reverse the movement
        stock_item.current_quantity -= movement.quantity_change
        stock_item.save()

        movement.delete()
        return redirect('stock_movement_list')

    return render(request, 'core/confirm_delete.html', {
        'title': 'Lagerabweichung Löschen',
        'message': f'Möchten Sie die Bewegung wirklich löschen? "{movement.stock_item.item_name}"?'
    })


@login_required
def export_invoices(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = [
        "Rechnungs Nr.",
        "Lieferant",
        "Rechnungsdatum",
        "Status",
        "Artikel",
        "Lieferanzahl",
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    invoices = Invoice.objects.prefetch_related("items", "delivery_notes").all()

    for invoice in invoices:
        items = ", ".join([
            f"{item.item_name} ({item.quantity} {item.unit})"
            for item in invoice.items.all()
        ])

        ws.append([
            invoice.invoice_number,
            invoice.supplier_name,
            invoice.invoice_date.strftime("%d.%m.%Y"),
            invoice.get_status_display(),
            items,
            invoice.delivery_notes.count(),
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="invoices.xlsx"'

    wb.save(response)
    return response


@login_required
def export_stock(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lager"

    headers = [
        "Artikel",
        "Aktuelle Menge",
        "Aktualisiert am",
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    stock_items = StockItem.objects.all().order_by("item_name")

    for item in stock_items:
        ws.append([
            item.item_name,
            item.current_quantity,
            item.updated_at.strftime("%d.%m.%Y %H:%M"),
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="stock.xlsx"'

    wb.save(response)
    return response


